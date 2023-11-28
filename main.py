from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.wait import WebDriverWait
import pyscreenshot as ImageGrab
import shutil
import openpyxl
import time
import os
import subprocess
import concurrent.futures

def downloadPrt():
    download_directory="E:\\001_QA_Generator\\downloaded_files\\prt_files"
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
    "download.default_directory":download_directory,
    "download.prompt_for_download": False, #To auto download the file
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
    })
    driver = webdriver.Chrome("E:\\001_QA_Generator\\chromedriver_win32\\chromedriver.exe",options=options)

    driver.get('https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id=38YN54')
    driver.maximize_window()
    driver.implicitly_wait(5)
    email=driver.find_element(By.ID,"cds-login-email")
    email.send_keys("c-skumar6@cdsvisual.com")
    password=driver.find_element(By.ID,"cds-login-passwd")
    password.send_keys("R7GXgpUASEwU")
    login_button=driver.find_element(By.TAG_NAME,'button').click()
    driver.implicitly_wait(5)

    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    i=1
    while(i<=row_count):
        part_number=load_ws.cell(row=i,column=1).value
        driver.get("https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id="+str(part_number)) 
        driver.implicitly_wait(10)    
        cad = driver.find_element(By.ID,"cds-cad-download-formats")
        select_cad=Select(cad)
        select_cad.select_by_index(5)
        time.sleep(1)
        driver.find_element(By.ID,"cds-cad-download-button").click()
        driver.implicitly_wait(200)
        time.sleep(1)
        driver.find_element(By.XPATH,('//*[@id="cds-cad-request-dialog"]/table/tbody/tr/td[2]/div[2]')).click()
        wait = WebDriverWait(driver, 100)
        wait.until(lambda x: any(filename.endswith('.zip')
            for filename in os.listdir(download_directory)))
        time.sleep(1)
        file_to_copy = 'E:/001_QA_Generator/downloaded_files/prt_files/grainger-'+part_number+'.zip'
        destination_directory = 'E:/001_QA_Generator/snap_tool/input/prt'
        shutil.move(file_to_copy, destination_directory)
        shutil.unpack_archive('E:/001_QA_Generator/snap_tool/input/prt/grainger-'+part_number+'.zip', 'E:/001_QA_Generator/snap_tool/input/prt/'+part_number+'_prt')
        os.remove('E:/001_QA_Generator/snap_tool/input/prt/grainger-'+part_number+'.zip')
        i+=1
    time.sleep(2)
    driver.close()

def downloadSimp():
    download_directory="E:\\001_QA_Generator\\downloaded_files\\simp_files"
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
    "download.default_directory":download_directory,
    "download.prompt_for_download": False, #To auto download the file
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
    })
    driver = webdriver.Chrome("E:\\001_QA_Generator\\chromedriver_win32\\chromedriver.exe",options=options)

    driver.get('https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id=38YN54')
    driver.maximize_window()
    driver.implicitly_wait(5)
    email=driver.find_element(By.ID,"cds-login-email")
    email.send_keys("c-skumar6@cdsvisual.com")
    password=driver.find_element(By.ID,"cds-login-passwd")
    password.send_keys("R7GXgpUASEwU")
    login_button=driver.find_element(By.TAG_NAME,'button').click()
    driver.implicitly_wait(5)

    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    i=1
    while(i<=row_count):
        part_number=load_ws.cell(row=i,column=1).value
        driver.get("https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id="+str(part_number))        
        cad = driver.find_element(By.ID,"cds-cad-download-formats")
        select_cad=Select(cad)
        select_cad.select_by_index(3)
        time.sleep(1)
        driver.find_element(By.ID,"cds-cad-download-button").click()
        driver.implicitly_wait(200)
        time.sleep(1)
        driver.find_element(By.XPATH,('//*[@id="cds-cad-request-dialog"]/table/tbody/tr/td[2]/div[2]')).click()
        wait = WebDriverWait(driver, 200)
        wait.until(lambda x: any(filename.endswith('.zip')
        for filename in os.listdir(download_directory)))
        time.sleep(1)
        file_to_copy = 'E:/001_QA_Generator/downloaded_files/simp_files/grainger-'+part_number+'.zip'
        destination_directory = 'E:/001_QA_Generator/snap_tool/input/stp_simp'
        shutil.move(file_to_copy, destination_directory)
        shutil.unpack_archive('E:/001_QA_Generator/snap_tool/input/stp_simp/grainger-'+part_number+'.zip', 'E:/001_QA_Generator/snap_tool/input/stp_simp/'+part_number+'_simp')
        os.remove('E:/001_QA_Generator/snap_tool/input/stp_simp/grainger-'+part_number+'.zip')
        i+=1
    time.sleep(1)
    driver.close()

def downloadSldprt():
    download_directory="E:\\001_QA_Generator\\downloaded_files\\sldprt_files"
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
    "download.default_directory":download_directory,
    "download.prompt_for_download": False, #To auto download the file
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
    })
    driver = webdriver.Chrome("E:\\001_QA_Generator\\chromedriver_win32\\chromedriver.exe",options=options)

    driver.get('https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id=38YN54')
    driver.maximize_window()
    driver.implicitly_wait(5)
    email=driver.find_element(By.ID,"cds-login-email")
    email.send_keys("c-skumar6@cdsvisual.com")
    password=driver.find_element(By.ID,"cds-login-passwd")
    password.send_keys("R7GXgpUASEwU")
    login_button=driver.find_element(By.TAG_NAME,'button').click()
    driver.implicitly_wait(5)

    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    i=1
    while(i<=row_count):
        part_number=load_ws.cell(row=i,column=1).value
        driver.get("https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id="+str(part_number))
        cad = driver.find_element(By.ID,"cds-cad-download-formats")
        select_cad=Select(cad)
        select_cad.select_by_index(4)
        time.sleep(1)
        driver.find_element(By.ID,"cds-cad-download-button").click()
        driver.implicitly_wait(200)
        time.sleep(1)
        driver.find_element(By.XPATH,('//*[@id="cds-cad-request-dialog"]/table/tbody/tr/td[2]/div[2]')).click()
        wait = WebDriverWait(driver, 200)
        wait.until(lambda x: any(filename.endswith('.zip')
            for filename in os.listdir(download_directory)))
        time.sleep(1)
        file_to_copy = 'E:/001_QA_Generator/downloaded_files/sldprt_files/grainger-'+part_number+'.zip'
        destination_directory = 'E:/001_QA_Generator/snap_tool/input/sldprt'
        shutil.move(file_to_copy, destination_directory)
        shutil.unpack_archive('E:/001_QA_Generator/snap_tool/input/sldprt/grainger-'+part_number+'.zip', 'E:/001_QA_Generator/snap_tool/input/sldprt/'+part_number+'_sldprt')
        os.remove('E:/001_QA_Generator/snap_tool/input/sldprt/grainger-'+part_number+'.zip')
        i+=1
    time.sleep(1)
    driver.close()

def downloadpdf():
    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    download_directory="E:\\001_QA_Generator\\downloaded_files\\downloaded_pdf"
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
    "download.default_directory":download_directory,
    "download.prompt_for_download": False, #To auto download the file
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
    })
    driver = webdriver.Chrome("E:\\001_QA_Generator\\chromedriver_win32\\chromedriver.exe",options=options)

    i=1
    driver.get('https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id=38YN54')
    driver.maximize_window()
    driver.implicitly_wait(5)
    email=driver.find_element(By.ID,"cds-login-email")
    email.send_keys("c-skumar6@cdsvisual.com")
    password=driver.find_element(By.ID,"cds-login-passwd")
    password.send_keys("R7GXgpUASEwU")
    login_button=driver.find_element(By.TAG_NAME,'button').click()
    driver.implicitly_wait(5)
    while(i<=row_count):
        part_number=load_ws.cell(row=i,column=1).value
        driver.get("https://qa.product-config.net/catalog3/d/grainger/?c=products&cid=root&id="+str(part_number))
        cad = driver.find_element(By.ID,"cds-cad-download-formats")
        select_cad=Select(cad)
        select_cad.select_by_index(1)
        driver.implicitly_wait(5)
        download_cad=driver.find_element(By.ID,"cds-cad-download-button")
        download_cad.click()
        driver.implicitly_wait(200)
        time.sleep(2)
        driver.find_element(By.XPATH,"/html/body/div[4]/div[2]/div[2]/a").click()
        driver.implicitly_wait(5)
        driver.find_element(By.XPATH,"/html/body/div[4]/div[1]/button/span[1]").click()
        time.sleep(3)
        im=ImageGrab.grab(bbox=(110,350,655,820))
        directory=f"E:\\001_QA_Generator\\png_images\\{part_number}"
        if not os.path.exists(directory):
            os.mkdir(directory)
        snip_directory=f"E:\\001_QA_Generator\\png_images\\{part_number}\\"
        file_nmae=part_number+'_json'
        file_save_name=snip_directory+file_nmae+'.png'
        im.save(file_save_name)
        i+=1
    time.sleep(2)
    driver.close()


def copyPartNumber():
    excel="E:\\001_QA_Generator\\snap_tool\\convert.xlsx"
    sheet_name = 'part_number'
    workbook = openpyxl.load_workbook("E:\\001_QA_Generator\\snap_tool\\convert.xlsx")
    workbook.create_sheet(title=sheet_name)
    workbook.save(excel)

    source_file = "E:\\001_QA_Generator\\input.xlsx"
    source_sheet_name = 'Sheet1'
    source_column = 'A'

    destination_file = "E:\\001_QA_Generator\\snap_tool\\convert.xlsx"
    destination_sheet_name = 'part_number'
    destination_column = 'A'

    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet_name]


    destination_workbook = openpyxl.load_workbook(destination_file)
    destination_sheet = destination_workbook[destination_sheet_name]

    max_row = source_sheet.max_row

    for row_number in range(1, max_row + 1):
        source_cell_value = source_sheet[source_column + str(row_number)].value
        destination_sheet[destination_column + str(row_number)].value = source_cell_value

    destination_workbook.save(destination_file)

def getfilename():
    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    j=1
    while(j<=row_count):
        part_number=(str(load_ws.cell(row=j,column=1).value))+"_sldprt"
        product_id=str(load_ws.cell(row=j,column=1).value)
        directory = f"E://001_QA_Generator//snap_tool//input//sldprt//{part_number}"
        excel_file = "E://001_QA_Generator//snap_tool//convert.xlsx"

        file_names = []

        for file in os.listdir(directory):
            file_names.append(file)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.create_sheet(product_id)

        row = 1
        col = 1
        for name in file_names:
            ws.cell(row, col).value = name
            row += 1
        wb.save(excel_file)
        j+=1


    
def openPdfInPdfViewer():
    file_name = "E:\\001_QA_Generator\\input.xlsx"
    load_wb = load_workbook(file_name, data_only=True)
    load_ws = load_wb['Sheet1']
    row_count = load_ws.max_row

    j=1
    while(j<=row_count):
        part_number=str(load_ws.cell(row=j,column=1).value)
        path=f".\downloaded_files\downloaded_pdf\grainger-{part_number}.pdf"
        acrobat_path = "C:\Program Files (x86)\Adobe\Acrobat 2020\Acrobat\Acrobat.exe" 
        app=subprocess.Popen([acrobat_path, path], shell=True)
        im=ImageGrab.grab(bbox=(425,205,1470,980))
        directory=f"E:\\001_QA_Generator\\png_images\\{part_number}"
        if not os.path.exists(directory):
            os.mkdir(directory)
        snip_directory=f"E:\\001_QA_Generator\\png_images\\{part_number}\\"
        file_nmae=part_number+'_pdf'
        file_save_name=snip_directory+file_nmae+'.png'
        im.save(file_save_name)
        j=j+1
    time.sleep(2)

def saveExcel():
    subprocess.run("E:\\001_QA_Generator\\snap_tool\\save_excel.bat", shell=True)

def main():
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(func) for func in [downloadPrt, downloadSimp, downloadSldprt, downloadpdf]]  
    concurrent.futures.wait(futures)
    copyPartNumber()
    getfilename()
    openPdfInPdfViewer()
    saveExcel()
       
if __name__ == "__main__":
    main()
